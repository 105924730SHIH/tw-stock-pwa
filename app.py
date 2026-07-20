# -*- coding: utf-8 -*-
"""
台股技術分析報表產生器 + Email 寄送（Gradio 介面版｜Render 部署版）

這是原本的 Gradio 後端程式（未變更任何分析邏輯）。
本資料夾內的 index.html / manifest.json / sw.js / icons 為新增的
PWA 外殼，會以 iframe 方式包住這個 Gradio 服務，讓使用者能把它
「安裝」到手機或電腦桌面、取得應用程式圖示，並在離線時顯示提示畫面。

安裝套件：
    pip install -r requirements.txt

執行：
    python app.py

啟動後預設會在 http://localhost:7860 提供服務，
接著開啟 index.html（或部署到網頁伺服器後開啟該網址），
在 PWA 的「設定」中輸入 http://localhost:7860 即可連線使用。
"""

import os
import re
import base64
import traceback
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import requests
import yfinance as yf
import resend
import gradio as gr

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference


# ══════════════════════════════════════════════════════════════════════════
# 輔助：抓取台股中文名稱對照表
# ══════════════════════════════════════════════════════════════════════════
def build_tw_name_map():
    urls = [
        "https://isin.twse.com.tw/isin/C_public.jsp?strMode=2",
        "https://isin.twse.com.tw/isin/C_public.jsp?strMode=4",
    ]
    name_map = {}
    headers = {"User-Agent": "Mozilla/5.0"}
    for u in urls:
        try:
            resp = requests.get(u, headers=headers, timeout=15)
            resp.encoding = "big5"
            tables = pd.read_html(resp.text)
            df = tables[0]
            df.columns = df.iloc[0]
            df = df.iloc[1:]
            col = df.columns[0]
            for val in df[col].dropna():
                val = str(val).strip()
                m = re.match(r"^(\d{4,6})\s*[\u3000 ]\s*(.+)$", val)
                if m:
                    code, name = m.group(1), m.group(2).strip()
                    if code not in name_map:
                        name_map[code] = name
        except Exception:
            pass
    return name_map


def get_display_name(ticker, tw_name_map):
    code = ticker.replace(".TWO", "").replace(".TW", "")
    name = tw_name_map.get(code)
    if name:
        return f"{ticker} {name}"
    try:
        info = yf.Ticker(ticker).info
        name = info.get("longName") or info.get("shortName") or ""
        if name:
            return f"{ticker} {name}"
    except Exception:
        pass
    return ticker


# ══════════════════════════════════════════════════════════════════════════
# 樣式
# ══════════════════════════════════════════════════════════════════════════
def make_styles():
    return {
        "header_font":  Font(name="Arial", bold=True, color="FFFFFF", size=11),
        "header_fill":  PatternFill("solid", start_color="0070C0"),
        "data_font":    Font(name="Arial", size=10),
        "alt_fill":     PatternFill("solid", start_color="EBF3FB"),
        "summary_fill": PatternFill("solid", start_color="D9E1F2"),
        "center":       Alignment(horizontal="center", vertical="center"),
        "right":        Alignment(horizontal="right",  vertical="center"),
        "border":       Border(left=Side(style="thin", color="BFBFBF"),
                                right=Side(style="thin", color="BFBFBF"),
                                top=Side(style="thin", color="BFBFBF"),
                                bottom=Side(style="thin", color="BFBFBF")),
    }


def write_price_sheet(ws, df, s, display_name):
    headers = ["日期", "開盤價 (TWD)", "最高價 (TWD)", "最低價 (TWD)",
               "收盤價 (TWD)", "成交量", "線性趨勢 (TWD)"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = s["header_font"]; cell.fill = s["header_fill"]
        cell.alignment = s["center"]; cell.border = s["border"]
    ws.row_dimensions[1].height = 20

    close_vals = df["收盤價"].values.astype(float)
    x = np.arange(len(close_vals))
    coeffs = np.polyfit(x, close_vals, 1)
    trend_vals = np.polyval(coeffs, x)

    for i, (_, row) in enumerate(df.iterrows()):
        ws.append(list(row) + [round(float(trend_vals[i]), 2)])
        er = ws.max_row
        for ci in range(1, 8):
            cell = ws.cell(row=er, column=ci)
            cell.font = s["data_font"]; cell.border = s["border"]
            cell.alignment = s["center"] if ci == 1 else s["right"]
            if er % 2 == 0:
                cell.fill = s["alt_fill"]
            if ci in (2, 3, 4, 5, 7):
                cell.number_format = "#,##0.00"
            elif ci == 6:
                cell.number_format = "#,##0"

    last_row = ws.max_row
    ss = last_row + 2
    title_c = ws.cell(row=ss, column=1, value="統計摘要")
    title_c.font = Font(name="Arial", bold=True, size=11, color="0070C0")

    for i, (lbl, fml, fmt) in enumerate(zip(
        ["最高收盤價", "最低收盤價", "平均收盤價", "期間漲跌幅"],
        [f"=MAX(E2:E{last_row})", f"=MIN(E2:E{last_row})",
         f"=AVERAGE(E2:E{last_row})", f"=(E{last_row}-E2)/E2"],
        ["#,##0.00", "#,##0.00", "#,##0.00", "0.00%"]
    )):
        r = ss + 1 + i
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(name="Arial", bold=True, size=10)
        lc.alignment = s["center"]; lc.border = s["border"]; lc.fill = s["summary_fill"]
        vc = ws.cell(row=r, column=2, value=fml)
        vc.font = s["data_font"]; vc.alignment = s["right"]
        vc.border = s["border"]; vc.number_format = fmt

    for i, w in enumerate([14, 16, 16, 16, 16, 16, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    chart = LineChart()
    chart.title = f"{display_name} 收盤價走勢"; chart.style = 10
    chart.y_axis.title = "收盤價 (TWD)"; chart.x_axis.title = "日期"
    chart.height = 15; chart.width = 28

    chart.add_data(Reference(ws, min_col=5, min_row=1, max_row=last_row), titles_from_data=True)
    chart.add_data(Reference(ws, min_col=7, min_row=1, max_row=last_row), titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last_row))

    c0 = chart.series[0]
    c0.graphicalProperties.line.solidFill = "0070C0"
    c0.graphicalProperties.line.width = 20000; c0.smooth = True

    c1 = chart.series[1]
    c1.graphicalProperties.line.solidFill = "FF4500"
    c1.graphicalProperties.line.width = 16000
    c1.graphicalProperties.line.dashDot = "dash"; c1.smooth = False

    ws.add_chart(chart, f"D{ss}")


# ══════════════════════════════════════════════════════════════════════════
# 技術指標計算
# ══════════════════════════════════════════════════════════════════════════
def _rsi(c, p=14):
    d = c.diff()
    ag = d.clip(lower=0).ewm(com=p - 1, min_periods=p).mean()
    al = (-d.clip(upper=0)).ewm(com=p - 1, min_periods=p).mean()
    return 100 - 100 / (1 + ag / al)


def _macd(c, f=12, s=26, sig=9):
    m = c.ewm(span=f, adjust=False).mean() - c.ewm(span=s, adjust=False).mean()
    sl = m.ewm(span=sig, adjust=False).mean()
    return m, sl, m - sl


def _kd(h, l, c, n=9, m=3):
    rsv = (c - l.rolling(n).min()) / (h.rolling(n).max() - l.rolling(n).min()) * 100
    K = rsv.ewm(com=m - 1, adjust=False).mean()
    return K, K.ewm(com=m - 1, adjust=False).mean()


def _obv(c, v):
    return (np.sign(c.diff()).fillna(0) * v).cumsum()


def _bollinger(c, p=20, k=2):
    m = c.rolling(p).mean(); sd = c.rolling(p).std()
    return m + k * sd, m, m - k * sd


def _atr(h, l, c, p=14):
    tr = pd.concat([h - l, (h - c.shift()).abs(), (l - c.shift()).abs()], axis=1).max(axis=1)
    return tr.ewm(com=p - 1, min_periods=p).mean()


def _adx(h, l, c, p=14):
    tr = pd.concat([h - l, (h - c.shift()).abs(), (l - c.shift()).abs()], axis=1).max(axis=1)
    up, dn = h.diff(), -l.diff()
    pdm = np.where((up > dn) & (up > 0), up.values, 0.0)
    ndm = np.where((dn > up) & (dn > 0), dn.values, 0.0)
    atr_s = tr.ewm(com=p - 1, min_periods=p).mean()
    pdi = 100 * pd.Series(pdm, index=c.index).ewm(com=p - 1, min_periods=p).mean() / atr_s
    ndi = 100 * pd.Series(ndm, index=c.index).ewm(com=p - 1, min_periods=p).mean() / atr_s
    dx = (abs(pdi - ndi) / (pdi + ndi) * 100).fillna(0)
    return dx.ewm(com=p - 1, min_periods=p).mean(), pdi, ndi


def _cci(h, l, c, p=20):
    tp = (h + l + c) / 3
    return (tp - tp.rolling(p).mean()) / (
        0.015 * tp.rolling(p).apply(lambda x: np.mean(np.abs(x - x.mean())), raw=True))


def _sar(h, l, af_step=0.02, af_max=0.2):
    n = len(h)
    sar = np.zeros(n); ep = np.zeros(n); af = np.zeros(n)
    bull = True
    sar[0] = float(l.iloc[0]); ep[0] = float(h.iloc[0]); af[0] = af_step
    for i in range(1, n):
        ps, pe, paf = sar[i - 1], ep[i - 1], af[i - 1]
        hi, lo = float(h.iloc[i]), float(l.iloc[i])
        h1, l1 = float(h.iloc[i - 1]), float(l.iloc[i - 1])
        h2, l2 = float(h.iloc[max(0, i - 2)]), float(l.iloc[max(0, i - 2)])
        if bull:
            s = min(ps + paf * (pe - ps), l1, l2)
            if lo < s:
                bull = False; sar[i] = pe; ep[i] = lo; af[i] = af_step
            else:
                sar[i] = s; ep[i] = max(pe, hi)
                af[i] = min(paf + af_step, af_max) if hi > pe else paf
        else:
            s = max(ps + paf * (pe - ps), h1, h2)
            if hi > s:
                bull = True; sar[i] = pe; ep[i] = hi; af[i] = af_step
            else:
                sar[i] = s; ep[i] = min(pe, lo)
                af[i] = min(paf + af_step, af_max) if lo < pe else paf
    return pd.Series(sar, index=h.index)


def _ichimoku(h, l, c):
    tk = (h.rolling(9).max() + l.rolling(9).min()) / 2
    kj = (h.rolling(26).max() + l.rolling(26).min()) / 2
    return (tk, kj,
            ((tk + kj) / 2).shift(26),
            ((h.rolling(52).max() + l.rolling(52).min()) / 2).shift(26),
            c.shift(-26))


def compute_indicators(df):
    c = df["收盤價"].astype(float)
    h = df["最高價"].astype(float)
    l = df["最低價"].astype(float)
    v = df["成交量"].astype(float)

    ind = pd.DataFrame({"日期": df["日期"].values, "收盤價": c.values}, index=df.index)

    ind["RSI(14)"] = _rsi(c).round(2)

    ml, ms, mh = _macd(c)
    ind["MACD"] = ml.round(4); ind["MACD_Signal"] = ms.round(4); ind["MACD_Hist"] = mh.round(4)

    ind["SMA20"] = c.rolling(20).mean().round(2)
    ind["SMA50"] = c.rolling(50).mean().round(2)
    ind["SMA200"] = c.rolling(200).mean().round(2)
    ind["EMA12"] = c.ewm(span=12, adjust=False).mean().round(2)
    ind["EMA26"] = c.ewm(span=26, adjust=False).mean().round(2)

    K, D = _kd(h, l, c)
    ind["K(9)"] = K.round(2); ind["D(3)"] = D.round(2)

    ind["OBV"] = _obv(c, v).round(0)

    bu, bm, bl = _bollinger(c)
    ind["BB_Upper"] = bu.round(2); ind["BB_Mid"] = bm.round(2); ind["BB_Lower"] = bl.round(2)
    ind["BBW"] = ((bu - bl) / bm).round(4)
    ind["%B"] = ((c - bl) / (bu - bl)).round(4)

    ind["BIAS(20)"] = ((c - bm) / bm * 100).round(2)
    ind["ROC(12)"] = ((c - c.shift(12)) / c.shift(12) * 100).round(2)
    ind["ATR(14)"] = _atr(h, l, c).round(4)

    adx_v, pdi, ndi = _adx(h, l, c)
    ind["ADX(14)"] = adx_v.round(2); ind["+DI"] = pdi.round(2); ind["-DI"] = ndi.round(2)

    ind["CCI(20)"] = _cci(h, l, c).round(2)
    ind["Momentum(10)"] = (c - c.shift(10)).round(2)
    ind["VWAP"] = ((c * v).cumsum() / v.cumsum()).round(2)
    ind["SAR"] = _sar(h, l).round(2)

    tk, kj, spa, spb, ck = _ichimoku(h, l, c)
    ind["Ichi_Tenkan"] = tk.round(2); ind["Ichi_Kijun"] = kj.round(2)
    ind["Ichi_SenkouA"] = spa.round(2); ind["Ichi_SenkouB"] = spb.round(2)
    ind["Ichi_Chikou"] = ck.round(2)

    ph, plv = float(h.max()), float(l.min())
    d = ph - plv
    for pct, ratio in [("0%", 0), ("23.6%", 0.236), ("38.2%", 0.382),
                        ("50%", 0.5), ("61.8%", 0.618), ("100%", 1.0)]:
        ind[f"Fib_{pct}"] = round(plv + ratio * d, 2)

    ind["Envelope_U"] = (bm * 1.025).round(2)
    ind["Envelope_L"] = (bm * 0.975).round(2)

    ind["Golden_Cross"] = np.where(
        ind["SMA50"].notna() & ind["SMA200"].notna() & (ind["SMA50"] > ind["SMA200"]),
        "黃金交叉(多)",
        np.where(ind["SMA50"].notna() & ind["SMA200"].notna() & (ind["SMA50"] < ind["SMA200"]),
                 "死亡交叉(空)", "N/A")
    )
    return ind


_FILL_Y = PatternFill("solid", start_color="FFFF00")
_FILL_R = PatternFill("solid", start_color="FF4444")
_FONT_K = Font(name="Arial", size=9, color="000000")
_FONT_W = Font(name="Arial", size=9, color="FFFFFF")
_FONT_N = Font(name="Arial", size=9)


def _get_color(col, val):
    if pd.isna(val):
        return None, None
    try:
        v = float(val)
    except Exception:
        return None, None

    if col == "RSI(14)":
        if v < 30: return _FILL_Y, _FONT_K
        if v > 70: return _FILL_R, _FONT_W
    elif col == "%B":
        if v < 0: return _FILL_Y, _FONT_K
        if v > 1: return _FILL_R, _FONT_W
    elif col == "ADX(14)":
        if v > 25: return _FILL_Y, _FONT_K
        if v < 20: return _FILL_R, _FONT_W
    elif col == "CCI(20)":
        if v < -100: return _FILL_Y, _FONT_K
        if v > 100: return _FILL_R, _FONT_W
    return None, None


def write_indicator_sheet(ws, df_ind, s):
    thin = Side(style="thin", color="BFBFBF")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    cols = list(df_ind.columns)

    ws.append(cols)
    for cell in ws[1]:
        cell.font = s["header_font"]; cell.fill = s["header_fill"]
        cell.alignment = s["center"]; cell.border = bdr
    ws.row_dimensions[1].height = 20

    alt = PatternFill("solid", start_color="EBF3FB")

    for ri, (_, row) in enumerate(df_ind.iterrows(), start=2):
        for ci, col in enumerate(cols, start=1):
            raw_val = row[col]
            if isinstance(raw_val, np.integer):
                val = int(raw_val)
            elif isinstance(raw_val, np.floating):
                val = None if np.isnan(raw_val) else float(raw_val)
            elif isinstance(raw_val, float) and np.isnan(raw_val):
                val = None
            else:
                val = raw_val

            cell = ws.cell(row=ri, column=ci, value=val)
            fill, font = _get_color(col, val)

            if fill:
                cell.fill = fill; cell.font = font
            else:
                cell.font = _FONT_N
                if ri % 2 == 0:
                    cell.fill = alt

            cell.border = bdr
            cell.alignment = Alignment(
                horizontal="center" if ci == 1 else "right", vertical="center")

            if col == "OBV" and val is not None:
                cell.number_format = "#,##0"
            elif isinstance(val, float):
                cell.number_format = "#,##0.00"

    for ci, col in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(col) + 2, 10), 22)
    ws.freeze_panes = "A2"


def generate_advice(ticker, df_ind, display_name):
    non_na = df_ind.dropna(subset=["RSI(14)"])
    last = non_na.iloc[-1] if not non_na.empty else df_ind.iloc[-1]

    def get(col, default=np.nan):
        try:
            return float(last[col]) if not pd.isna(last[col]) else default
        except Exception:
            return default

    buy_signals = []
    sell_signals = []

    rsi = get("RSI(14)")
    macd = get("MACD")
    macd_sig = get("MACD_Signal")
    macd_h = get("MACD_Hist")
    k = get("K(9)")
    d = get("D(3)")
    pct_b = get("%B")
    adx = get("ADX(14)")
    pdi = get("+DI")
    ndi = get("-DI")
    cci = get("CCI(20)")
    close = get("收盤價")
    sar = get("SAR")
    bias = get("BIAS(20)")
    mom = get("Momentum(10)")
    golden = str(last.get("Golden_Cross", "N/A"))

    if not np.isnan(rsi):
        if rsi < 30: buy_signals.append(f"RSI={rsi:.1f}（超賣區 <30，潛在反彈機會）")
        elif 30 <= rsi <= 50: buy_signals.append(f"RSI={rsi:.1f}（中性偏低，動能回升中）")

    if not np.isnan(macd) and not np.isnan(macd_sig):
        if macd > macd_sig and not np.isnan(macd_h) and macd_h > 0:
            buy_signals.append(f"MACD 柱狀圖翻正（{macd_h:.4f}），多頭動能增強")

    if not np.isnan(k) and not np.isnan(d):
        if k < 30 and d < 30: buy_signals.append(f"KD={k:.1f}/{d:.1f}（雙雙低於30，超賣訊號）")
        elif k > d and k < 50: buy_signals.append(f"KD={k:.1f}/{d:.1f}（K 上穿 D 且位於低檔，黃金交叉）")

    if not np.isnan(pct_b) and pct_b < 0.2:
        buy_signals.append(f"%B={pct_b:.2f}（接近布林下緣，超賣支撐）")
    if not np.isnan(cci) and cci < -100:
        buy_signals.append(f"CCI={cci:.1f}（低於 -100，超賣反彈訊號）")
    if not np.isnan(close) and not np.isnan(sar) and close > sar:
        buy_signals.append(f"收盤({close:.2f}) > SAR({sar:.2f})，SAR 多頭排列")
    if golden == "黃金交叉(多)":
        buy_signals.append("SMA50 > SMA200，長線黃金交叉，趨勢向上")
    if not np.isnan(adx) and not np.isnan(pdi) and not np.isnan(ndi):
        if adx > 25 and pdi > ndi:
            buy_signals.append(f"ADX={adx:.1f}（趨勢強勁）且 +DI > -DI，多頭趨勢確立")
    if not np.isnan(mom) and mom > 0:
        buy_signals.append(f"10日動能={mom:.2f}（正值，短期上升動能）")

    if not np.isnan(rsi):
        if rsi > 70: sell_signals.append(f"RSI={rsi:.1f}（超買區 >70，短線獲利了結風險）")
        elif 50 < rsi <= 70: sell_signals.append(f"RSI={rsi:.1f}（偏高區間，留意回調）")

    if not np.isnan(macd) and not np.isnan(macd_sig):
        if macd < macd_sig and not np.isnan(macd_h) and macd_h < 0:
            sell_signals.append(f"MACD 柱狀圖翻負（{macd_h:.4f}），空頭動能增強")

    if not np.isnan(k) and not np.isnan(d):
        if k > 80 and d > 80: sell_signals.append(f"KD={k:.1f}/{d:.1f}（雙雙高於80，超買訊號）")
        elif k < d and k > 50: sell_signals.append(f"KD={k:.1f}/{d:.1f}（K 下穿 D 且位於高檔，死亡交叉）")

    if not np.isnan(pct_b) and pct_b > 0.8:
        sell_signals.append(f"%B={pct_b:.2f}（接近布林上緣，超買壓力）")
    if not np.isnan(cci) and cci > 100:
        sell_signals.append(f"CCI={cci:.1f}（超過 +100，超買回落訊號）")
    if not np.isnan(close) and not np.isnan(sar) and close < sar:
        sell_signals.append(f"收盤({close:.2f}) < SAR({sar:.2f})，SAR 空頭排列")
    if golden == "死亡交叉(空)":
        sell_signals.append("SMA50 < SMA200，長線死亡交叉，趨勢向下")
    if not np.isnan(adx) and not np.isnan(pdi) and not np.isnan(ndi):
        if adx > 25 and ndi > pdi:
            sell_signals.append(f"ADX={adx:.1f}（趨勢強勁）且 -DI > +DI，空頭趨勢確立")
    if not np.isnan(bias) and bias > 10:
        sell_signals.append(f"BIAS={bias:.1f}%（乖離率過高，均值回歸壓力）")
    if not np.isnan(mom) and mom < 0:
        sell_signals.append(f"10日動能={mom:.2f}（負值，短期下跌動能）")

    b, sv = len(buy_signals), len(sell_signals)
    net = b - sv

    if net >= 4: action = "強力買進 ★★★"; confidence = "高"
    elif net >= 2: action = "買進 ★★"; confidence = "中"
    elif net == 1: action = "觀望偏多 ★"; confidence = "低"
    elif net == -1: action = "觀望偏空 ▼"; confidence = "低"
    elif net <= -4: action = "強力賣出 ▼▼▼"; confidence = "高"
    elif net <= -2: action = "賣出 ▼▼"; confidence = "中"
    else: action = "觀望（中性）"; confidence = "低"

    reasons_buy = "；".join(buy_signals) if buy_signals else "無明顯買進訊號"
    reasons_sell = "；".join(sell_signals) if sell_signals else "無明顯賣出訊號"

    key_summary = (
        f"RSI={rsi:.1f}｜MACD柱={macd_h:.3f}｜KD={k:.1f}/{d:.1f}｜"
        f"%B={pct_b:.2f}｜ADX={adx:.1f}｜CCI={cci:.1f}｜"
        f"趨勢={golden}"
    )

    return {
        "建議": action,
        "信心": confidence,
        "買進訊號": reasons_buy,
        "賣出訊號": reasons_sell,
        "關鍵指標": key_summary,
        "公司名稱": display_name,
    }


_FILL_BUY3 = PatternFill("solid", start_color="00B050")
_FILL_BUY2 = PatternFill("solid", start_color="92D050")
_FILL_BUY1 = PatternFill("solid", start_color="E2EFDA")
_FILL_NEUT = PatternFill("solid", start_color="F2F2F2")
_FILL_SELL1 = PatternFill("solid", start_color="FFDDC1")
_FILL_SELL2 = PatternFill("solid", start_color="FF7070")
_FILL_SELL3 = PatternFill("solid", start_color="C00000")


def _advice_fill(action):
    if "強力買進" in action: return _FILL_BUY3, Font(name="Arial", bold=True, size=10, color="FFFFFF")
    if "買進 ★★" in action: return _FILL_BUY2, Font(name="Arial", bold=True, size=10, color="000000")
    if "偏多" in action: return _FILL_BUY1, Font(name="Arial", size=10, color="000000")
    if "強力賣出" in action: return _FILL_SELL3, Font(name="Arial", bold=True, size=10, color="FFFFFF")
    if "賣出 ▼▼" in action: return _FILL_SELL2, Font(name="Arial", bold=True, size=10, color="FFFFFF")
    if "偏空" in action: return _FILL_SELL1, Font(name="Arial", size=10, color="000000")
    return _FILL_NEUT, Font(name="Arial", size=10, color="000000")


def write_advice_sheet(ws, all_advice, s):
    thin = Side(style="thin", color="BFBFBF")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["股票代號／公司名稱", "最新建議", "信心程度", "買進訊號說明", "賣出訊號說明", "關鍵指標摘要"]
    col_widths = [28, 20, 10, 60, 60, 70]

    ws.append(headers)
    for ci, cell in enumerate(ws[1], start=1):
        cell.font = s["header_font"]; cell.fill = s["header_fill"]
        cell.alignment = s["center"]; cell.border = bdr
    ws.row_dimensions[1].height = 22

    ws.append(["⬛ 色彩圖例",
               "深綠=強力買進 / 淺綠=買進 / 極淺綠=觀望偏多",
               "灰=中性觀望",
               "淺橙=觀望偏空 / 橙紅=賣出 / 深紅=強力賣出",
               "", "※ 本建議僅供參考，不構成投資依據"])
    for ci in range(1, 7):
        cell = ws.cell(row=2, column=ci)
        cell.font = Font(name="Arial", size=9, italic=True, color="595959")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bdr
    ws.row_dimensions[2].height = 30

    for ri, (ticker, adv) in enumerate(all_advice.items(), start=3):
        row_data = [
            adv["公司名稱"], adv["建議"], adv["信心"],
            adv["買進訊號"], adv["賣出訊號"], adv["關鍵指標"],
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = bdr
            cell.alignment = Alignment(
                horizontal="center" if ci <= 3 else "left",
                vertical="center", wrap_text=True
            )
            if ci == 2:
                fill, font = _advice_fill(val)
                cell.fill = fill; cell.font = font
            else:
                cell.font = Font(name="Arial", size=10)
                if ri % 2 == 0:
                    cell.fill = PatternFill("solid", start_color="F5F5F5")
        ws.row_dimensions[ri].height = 60

    for ci, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A3"

    ws.insert_rows(1)
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1,
                          value=f"台股投資建議彙整表｜分析日期：{datetime.today().strftime('%Y-%m-%d')}")
    title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    title_cell.fill = PatternFill("solid", start_color="1F497D")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32


# ══════════════════════════════════════════════════════════════════════════
# 主流程（供 Gradio 呼叫，使用 generator 逐步回報進度）
# ══════════════════════════════════════════════════════════════════════════
def run_pipeline(sheet_id, gid, resend_api_key, email_from, email_to, send_email):
    log_lines = []

    def log(msg):
        log_lines.append(msg)
        return "\n".join(log_lines)

    # ── 基本檢查 ──
    sheet_id = (sheet_id or "").strip()
    gid = (gid or "0").strip()
    resend_api_key = (resend_api_key or "").strip()
    email_from = (email_from or "").strip() or "Acme <onboarding@resend.dev>"
    email_to_list = [e.strip() for e in (email_to or "").split(",") if e.strip()]

    if not sheet_id:
        yield log("❌ 請輸入 Google Sheet ID"), None, _status("error", "缺少 Sheet ID")
        return
    if send_email and not resend_api_key:
        yield log("❌ 已勾選寄送 Email，但未輸入 Resend API Key"), None, _status("error", "缺少 API Key")
        return
    if send_email and not email_to_list:
        yield log("❌ 已勾選寄送 Email，但未輸入收件人信箱"), None, _status("error", "缺少收件人")
        return

    output_path = "tw_stocks_3months.xlsx"

    try:
        # ── 讀取股票清單 ──
        yield log("正在讀取 Google Sheet 股票清單..."), None, _status("busy", "讀取股票清單中")
        url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"
        df1 = pd.read_csv(url)
        if "TICKERS" not in df1.columns:
            yield log("❌ Google Sheet 中找不到 'TICKERS' 欄位，請確認格式"), None, _status("error", "欄位格式錯誤")
            return
        tickers = df1["TICKERS"].dropna().astype(str).str.strip().tolist()
        if not tickers:
            yield log("❌ 股票清單是空的"), None, _status("error", "清單是空的")
            return
        yield log(f"取得 {len(tickers)} 檔股票代號：{', '.join(tickers)}"), None, _status("busy", f"共 {len(tickers)} 檔待處理")

        # ── 中文名稱對照 ──
        yield log("正在抓取台股中文名稱對照表..."), None, _status("busy", "建立名稱對照表")
        tw_name_map = build_tw_name_map()
        yield log(f"取得 {len(tw_name_map)} 檔中文名稱"), None, _status("busy", "名稱對照完成")

        ticker_names = {}
        for t in tickers:
            ticker_names[t] = get_display_name(t, tw_name_map)

        end_date = datetime.today() + timedelta(days=1)
        start_date = end_date - timedelta(days=91)

        wb = Workbook()
        wb.remove(wb.active)
        s = make_styles()
        failed, success = [], []
        all_advice = {}

        yield log(f"下載區間：{start_date.date()} ~ {end_date.date()}"), None, _status("busy", "開始下載股價資料")

        for idx, ticker in enumerate(tickers, start=1):
            display_name = ticker_names.get(ticker, ticker)
            short = ticker.replace(".TWO", "").replace(".TW", "")
            progress_note = f"({idx}/{len(tickers)}) {display_name}"

            try:
                raw = yf.download(ticker, start=start_date, end=end_date,
                                   auto_adjust=True, progress=False)
                if raw.empty:
                    failed.append(ticker)
                    yield log(f"  ▶ {display_name} ... ❌ 無資料"), None, _status("busy", progress_note)
                    continue

                raw = raw.reset_index()
                raw.columns = [col[0] if isinstance(col, tuple) else col for col in raw.columns]
                raw["Date"] = pd.to_datetime(raw["Date"]).dt.date
                raw = raw[["Date", "Open", "High", "Low", "Close", "Volume"]]
                raw.columns = ["日期", "開盤價", "最高價", "最低價", "收盤價", "成交量"]
                raw = raw.sort_values("日期").reset_index(drop=True)

                ws_price = wb.create_sheet(title=short)
                write_price_sheet(ws_price, raw, s, display_name)

                df_ind = compute_indicators(raw)
                ws_ind = wb.create_sheet(title=(short + "_指標")[:31])
                write_indicator_sheet(ws_ind, df_ind, s)

                all_advice[ticker] = generate_advice(ticker, df_ind, display_name)

                success.append(ticker)
                yield log(f"  ▶ {display_name} ... ✅ {len(raw)} 筆"), None, _status("busy", progress_note)
            except Exception as e:
                failed.append(ticker)
                yield log(f"  ▶ {display_name} ... ❌ {e}"), None, _status("busy", progress_note)

        if all_advice:
            ws_adv = wb.create_sheet(title="投資建議彙整", index=0)
            write_advice_sheet(ws_adv, all_advice, s)

        wb.save(output_path)

        yield log(f"\n✅ Excel 產生完成！成功 {len(success)} 檔｜失敗 {len(failed)} 檔"), None, _status("busy", "Excel 產生完成")
        if failed:
            yield log(f"   失敗清單：{', '.join(failed)}"), None, _status("busy", "Excel 產生完成")

        # ── 寄送 Email（選用）──
        if send_email:
            yield log("\n正在寄送 Email..."), output_path, _status("busy", "寄送 Email 中")
            resend.api_key = resend_api_key

            with open(output_path, "rb") as f:
                file_content = base64.b64encode(f.read()).decode("utf-8")

            buy_top = [a["公司名稱"] for a in all_advice.values() if "買進" in a["建議"]][:10]
            sell_top = [a["公司名稱"] for a in all_advice.values() if "賣出" in a["建議"]][:10]

            html_body = f"""
            <h2>台股投資建議彙整表｜{datetime.today().strftime('%Y-%m-%d')}</h2>
            <p>成功分析 {len(success)} 檔股票，失敗 {len(failed)} 檔。</p>
            <p><b>買進訊號股票：</b>{', '.join(buy_top) if buy_top else '無'}</p>
            <p><b>賣出訊號股票：</b>{', '.join(sell_top) if sell_top else '無'}</p>
            <p>完整報表請見附件 Excel 檔案。</p>
            <p style="color:gray;font-size:12px;">※ 本郵件內容僅供參考，不構成投資建議。</p>
            """

            params = {
                "from": email_from,
                "to": email_to_list,
                "subject": f"台股交易報告 {datetime.today().strftime('%Y-%m-%d')}",
                "html": html_body,
                "attachments": [{
                    "filename": output_path,
                    "content": file_content,
                }],
            }

            email_result = resend.Emails.send(params)
            yield log(f"📧 郵件已寄出：{email_result}"), output_path, _status("done", f"完成｜成功 {len(success)}／失敗 {len(failed)}｜已寄出 Email")
        else:
            yield log("\n(未勾選寄送 Email，僅產生 Excel 檔案)"), output_path, _status("done", f"完成｜成功 {len(success)}／失敗 {len(failed)}")

    except Exception as e:
        yield log(f"❌ 發生錯誤：{e}\n{traceback.format_exc()}"), None, _status("error", "執行失敗")
        return


def _status(kind, text):
    """回傳一段狀態列 HTML，供 UI 顯示彩色狀態徽章。"""
    palette = {
        "idle":  ("#94a3b8", "#f1f5f9", "⏳"),
        "busy":  ("#2563eb", "#eff6ff", "🔄"),
        "done":  ("#16a34a", "#ecfdf5", "✅"),
        "error": ("#dc2626", "#fef2f2", "⚠️"),
    }
    color, bg, icon = palette.get(kind, palette["idle"])
    return (
        f'<div style="display:flex;align-items:center;gap:8px;padding:10px 14px;'
        f'border-radius:10px;background:{bg};color:{color};font-weight:600;'
        f'font-size:14px;border:1px solid {color}33;">'
        f'<span>{icon}</span><span>{text}</span></div>'
    )


# ══════════════════════════════════════════════════════════════════════════
# Gradio 介面（精緻化版）
# ══════════════════════════════════════════════════════════════════════════
CUSTOM_CSS = """
:root {
    --brand-blue: #1f497d;
    --brand-blue-light: #eff6ff;
}
.gradio-container {
    max-width: 1180px !important;
    margin: 0 auto !important;
}
#hero {
    background: linear-gradient(135deg, #1f497d 0%, #2563eb 60%, #0ea5e9 100%);
    border-radius: 18px;
    padding: 28px 32px;
    color: #ffffff !important;
    margin-bottom: 18px;
    box-shadow: 0 8px 24px rgba(31,73,125,0.25);
}
#hero * {
    color: #ffffff !important;
}
#hero h1 {
    margin: 0 0 6px 0;
    font-size: 26px;
    font-weight: 800;
    letter-spacing: 0.5px;
    color: #ffffff !important;
}
#hero p {
    margin: 0;
    opacity: 1;
    font-size: 14.5px;
    line-height: 1.6;
    color: #ffffff !important;
}
.panel-card {
    border-radius: 14px !important;
    border: 1px solid #e5e9f0 !important;
    box-shadow: 0 1px 3px rgba(15,23,42,0.06) !important;
}
.section-title {
    font-weight: 700 !important;
    font-size: 15px !important;
    color: #1f2937 !important;
    margin-bottom: 2px !important;
}
#run-btn {
    height: 48px !important;
    font-size: 16px !important;
    font-weight: 700 !important;
    border-radius: 12px !important;
    box-shadow: 0 4px 12px rgba(37,99,235,0.35) !important;
}
#log-box textarea {
    font-family: "SFMono-Regular", Consolas, Menlo, monospace !important;
    font-size: 12.5px !important;
    line-height: 1.5 !important;
    background: #0b1220 !important;
    color: #d1e2ff !important;
    border-radius: 12px !important;
}
footer {visibility: hidden}
"""

THEME = gr.themes.Soft(
    primary_hue="blue",
    secondary_hue="slate",
    neutral_hue="slate",
    font=[gr.themes.GoogleFont("Inter"), "Noto Sans TC", "sans-serif"],
).set(
    button_primary_background_fill="linear-gradient(90deg, #1f497d, #2563eb)",
    button_primary_background_fill_hover="linear-gradient(90deg, #17385f, #1d4ed8)",
    block_radius="14px",
    block_shadow="0 1px 3px rgba(15,23,42,0.06)",
)

with gr.Blocks(title="台股技術分析報表產生器") as demo:

    gr.HTML(
        """
        <div id="hero">
            <h1 style="color:#ffffff;">📈 台股技術分析報表產生器</h1>
            <p style="color:#ffffff;">
                從 Google Sheet 讀取股票代號 → 自動抓取近 3 個月股價、計算 20+ 項技術指標、
                產生買賣建議彙整表 → 匯出精美 Excel，並可選擇寄送 Email 通知。
            </p>
        </div>
        """
    )

    with gr.Row(equal_height=False):
        # ── 左側：設定區 ──
        with gr.Column(scale=4):
            with gr.Group(elem_classes="panel-card"):
                gr.Markdown("### 📊 資料來源設定", elem_classes="section-title")
                sheet_id_in = gr.Textbox(
                    label="Google Sheet ID",
                    placeholder="例如：1KXNmLdK9gL7QGaSYLaIAB5sq2TzKL65e7UBfJNK9xRo",
                    info="Sheet 需包含 TICKERS 欄位，列出股票代號（如 2330.TW）",
                )
                gid_in = gr.Textbox(
                    label="Sheet GID",
                    value="0",
                    placeholder="例如：0",
                    info="工作表分頁的 gid 參數，預設分頁通常為 0",
                )

            with gr.Accordion("📧 Email 寄送設定（選用）", open=True, elem_classes="panel-card"):
                send_email_chk = gr.Checkbox(label="產生報表後自動寄送 Email", value=True)
                resend_key_in = gr.Textbox(
                    label="Resend API Key",
                    type="password",
                    placeholder="re_xxxxxxxxxxxx",
                )
                email_from_in = gr.Textbox(
                    label="寄件者 (From)",
                    value="Acme <onboarding@resend.dev>",
                    placeholder="Name <you@yourdomain.com>",
                )
                email_to_in = gr.Textbox(
                    label="收件者 (To)",
                    placeholder="you@example.com, other@example.com",
                    info="多筆收件人請用逗號分隔",
                )

            run_btn = gr.Button("🚀 開始產生報表", variant="primary", elem_id="run-btn")

            status_out = gr.HTML(_status("idle", "尚未開始"))

        # ── 右側：執行紀錄 / 下載 ──
        with gr.Column(scale=6):
            with gr.Group(elem_classes="panel-card"):
                gr.Markdown("### 📋 執行紀錄", elem_classes="section-title")
                log_out = gr.Textbox(
                    label=None,
                    show_label=False,
                    lines=22,
                    max_lines=40,
                    elem_id="log-box",
                    placeholder="按下「開始產生報表」後，執行過程會即時顯示在這裡...",
                )
            with gr.Group(elem_classes="panel-card"):
                gr.Markdown("### 📥 報表下載", elem_classes="section-title")
                file_out = gr.File(label="Excel 報表（含技術指標與投資建議）")

    gr.Markdown(
        "<div style='text-align:center;color:#94a3b8;font-size:12px;margin-top:10px;'>"
        "⚠️ 本工具產出之內容僅供技術參考，不構成任何投資建議。"
        "</div>"
    )

    run_btn.click(
        fn=run_pipeline,
        inputs=[sheet_id_in, gid_in, resend_key_in, email_from_in, email_to_in, send_email_chk],
        outputs=[log_out, file_out, status_out],
    )

if __name__ == "__main__":
    # Render 會透過 PORT 環境變數指定公開連接埠；本機執行時則使用 7860。
    port = int(os.environ.get("PORT", "7860"))

    demo.queue().launch(
        server_name="0.0.0.0",
        server_port=port,
        theme=THEME,
        css=CUSTOM_CSS,
        show_error=True,
    )
